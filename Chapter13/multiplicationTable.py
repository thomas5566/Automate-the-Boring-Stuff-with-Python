"""
Creates a N x N multiplication table in a spreadsheet.
Usage: multiplication_table.py <N> - Creates N X N table
"""

import sys, openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# number = int(sys.argv[1]) # CMD input number
number = 9
cells = number + 1

wb = openpyxl.Workbook()
sheet = wb.active

make_bold = Font(bold=True)

while cells > 1:
    # Write outer cells values and make them bold
    sheet.cell(row=cells, column=1).value = cells - 1 # A10 = 9, A9 = 8 .... A2 = 1
    sheet.cell(row=1, column=cells).value = cells - 1 # J1 = 9, I1 = 8 .... B1 = 1

    sheet.cell(row=cells, column=1).font = make_bold
    sheet.cell(row=1, column=cells).font = make_bold

    cells -= 1

# Populate the table inners with correct formula
col_length = number + 1

count = 0

while count < number:
    col_letter = get_column_letter(sheet.max_column - count)

    while col_length > 1: # E.X: 9X9  J, I, H, G, F, E, D, C, B, A
        sheet[col_letter + str(col_length)] = ('=SUM(' + col_letter + '1*A' + str(col_length) + ')')
        # J10 = SUM(J1 * A10)
        #  J9 = SUM(J1 * A9)
        #  J8 = SUM(J1 * A8)
        #  J7 = SUM(J1 * A7)
        #  .....
        #  J2 = SUM(J1 * A1)
        col_length -= 1  # J10, J9, J8 .....

    col_length = number + 1  # Next Column
    count += 1


wb.save('multi_table.xlsx')




