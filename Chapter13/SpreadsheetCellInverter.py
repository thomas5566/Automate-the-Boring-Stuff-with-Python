import openpyxl
import re, os

print('Enter the absolute path to the spreadsheet file: ')
file_path = 'D:\\Users\\Administrator\\reverseFile.xlsx'

path_regex = re.compile(r'(.*)(\.xlsx)$')
path_split = path_regex.search(file_path)
path = path_split.group(1)
name = os.path.basename(file_path)

wb = openpyxl.load_workbook(file_path)
sheet = wb.active

max_row = sheet.max_row
max_column = sheet.max_column

# Make nested list of spreadsheet data (row by row)
rows = []
for row in range(1, max_row + 1):
    data = []
    for column in range(1, max_column + 1):
        cell_value = sheet.cell(row=row, column=column).value
        data.append(cell_value)
    rows.append(data)


# Open new spreadsheet and populate with above data but inverted
wb = openpyxl.Workbook()
sheet = wb.active

column_num =1
for row in rows:
    row_num = 1 # reset row index
    for cell in row:
        sheet.cell(row=row_num, column=column_num).value = cell
        row_num += 1
    column_num += 1 # column 累加不 reset index

wb.save(path + name + '(inverted).xlsx')
print('Spreadsheet data inverted and saved to ' + 'New_' + name)
