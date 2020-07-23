"""Insert the contents of multiple text files into a single spreadsheet."""

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

print('Enter the ablsolute path to the folder where the file :')
path = input()

print('Enter the file names you wish to have inserted into the EXCEL .txt file only :')
files = input()
file_list = files.split()

wb = openpyxl.Workbook()
sheet = wb.active

column_num = 1
for file in file_list:

    # Open file and format lines
    lines = open(path + '\\' + file).readlines()

    make_bold = Font(bold=True)
    sheet.cell(row=1, column=column_num).value = file
    sheet.cell(row=1, column=column_num).font = make_bold

    longest = 0 # 檔案內字串的總長度
    row_num = 2 # 檔案名稱位於 row1 column1, 所以讀進來的內容從row2開始

    for line in lines:
        line = line.strip()

        # Calculate appropriate column width
        if len(line) > longest:
            longest = len(line)

        # Write lines to spreadsheet
        sheet.cell(row=row_num, column=column_num).value = line
        row_num += 1

    column_letter = get_column_letter(column_num) # 把抓到最常字串的長度, 設為column的寬度
    sheet.column_dimensions[column_letter].width = longest

wb.save(path + '/text2sheet.xlsx')

print('Spreadsheet saved as text2sheet.xlsx - it can be found in the same directory as the inputted files.')