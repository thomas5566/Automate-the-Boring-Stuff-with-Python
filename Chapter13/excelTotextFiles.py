import openpyxl, re

print('Enter the absolute path of the EXCEL file wish to TEXT file:')
# file_path = input()
file_path = 'D:\\Users\\Administrator\\produceSales.xlsx'
path_regex = re.compile(r'(.*\\)(.*)(\.xlsx)$')
path_split = path_regex.search(file_path)
path = path_split.group(1)
name = path_split.group(2)

wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

max_column = sheet.max_column
max_row = sheet.max_row


col_num = 1
for column in range(1, max_column + 1):

    col_data = []
    for cell in range(1, max_row + 1):
        if sheet.cell(row=cell, column=col_num).value != None:
            col_data.append(sheet.cell(row=cell, column=col_num).value)
    col_num += 1

    file = open(path + 'col-' + str(col_num) + '-' + name + '.txt', 'w')

    for line in col_data:
        file.write(str(line) + '\n')
    file.close()

